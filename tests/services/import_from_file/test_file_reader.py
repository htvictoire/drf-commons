"""
Tests for FileReader class.

Tests file reading functionality.
"""

import os
import tempfile
from unittest.mock import patch

import pandas as pd

from drf_commons.common_tests.base_cases import DrfCommonTestCase
from drf_commons.common_tests.utils import create_csv_file

from drf_commons.services.import_from_file.core.file_reader import FileReader


class FileReaderTests(DrfCommonTestCase):
    """Tests for FileReader."""

    def setUp(self):
        super().setUp()
        self.config = {"file_format": "csv"}
        self.reader = FileReader(self.config)

    def test_file_reader_initialization(self):
        """Test file reader initializes correctly."""
        config = {"file_format": "csv"}
        reader = FileReader(config)
        self.assertIsInstance(reader, FileReader)

    def test_read_csv_file(self):
        """Test reading CSV file."""
        # Create test CSV file
        headers = ["username", "email"]
        rows = [["testuser1", "test1@example.com"], ["testuser2", "test2@example.com"]]
        csv_file = create_csv_file(headers, rows)

        # Create temporary file
        with tempfile.NamedTemporaryFile(
            mode="wb", delete=False, suffix=".csv"
        ) as temp_file:
            temp_file.write(csv_file.read())
            temp_file_path = temp_file.name

        try:
            df = self.reader.read_file(temp_file_path)
            self.assertIsInstance(df, pd.DataFrame)
            self.assertEqual(len(df), 2)
            self.assertIn("username", df.columns)
            self.assertIn("email", df.columns)
        finally:
            os.unlink(temp_file_path)

    def test_read_excel_file(self):
        """Test reading Excel file."""
        try:
            # Create reader with xlsx config
            excel_config = {"file_format": "xlsx"}
            excel_reader = FileReader(excel_config)

            # Create test Excel file with headers at row 5 (0-indexed row 4)
            import openpyxl
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # Add 4 empty rows first
            for i in range(1, 5):
                worksheet.cell(row=i, column=1, value="")

            # Write headers at row 5
            headers = ["username", "email"]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=5, column=col, value=header)

            # Write data rows
            rows = [
                ["testuser1", "test1@example.com"],
                ["testuser2", "test2@example.com"],
            ]
            for row_idx, row in enumerate(rows, 6):  # Start at row 6
                for col_idx, value in enumerate(row, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=value)

            # Create temporary file
            with tempfile.NamedTemporaryFile(
                mode="wb", delete=False, suffix=".xlsx"
            ) as temp_file:
                workbook.save(temp_file.name)
                temp_file_path = temp_file.name

            df = excel_reader.read_file(temp_file_path)
            self.assertIsInstance(df, pd.DataFrame)
            self.assertEqual(len(df), 2)
            self.assertIn("username", df.columns)
            self.assertIn("email", df.columns)

            os.unlink(temp_file_path)
        except ImportError:
            # Skip test if openpyxl not available
            self.skipTest("openpyxl not available for Excel file testing")

    def test_read_nonexistent_file(self):
        """Test reading nonexistent file raises appropriate error."""
        with self.assertRaises((FileNotFoundError, IOError)):
            self.reader.read_file("/nonexistent/file.csv")

    @patch("drf_commons.services.import_from_file.core.file_reader.pd.read_csv")
    def test_read_csv_calls_pandas_read_csv(self, mock_read_csv):
        """Test read_file calls pandas read_csv for CSV files."""
        mock_read_csv.return_value = pd.DataFrame()

        self.reader.read_file("test.csv")

        mock_read_csv.assert_called_once()

    @patch("drf_commons.services.import_from_file.core.file_reader.pd.read_excel")
    def test_read_xls_file_uses_xlrd_engine(self, mock_read_excel):
        """Test reading XLS file uses xlrd engine."""
        mock_read_excel.return_value = pd.DataFrame({"col": ["val"]})
        xls_config = {"file_format": "xls"}
        xls_reader = FileReader(xls_config)

        xls_reader.read_file("test.xls")

        mock_read_excel.assert_called_once()
        call_kwargs = mock_read_excel.call_args
        self.assertEqual(call_kwargs[1]["engine"], "xlrd")

    @patch("drf_commons.services.import_from_file.core.file_reader.pd.read_excel")
    def test_read_xlsx_file_uses_openpyxl_engine(self, mock_read_excel):
        """Test reading XLSX file uses openpyxl engine."""
        mock_read_excel.return_value = pd.DataFrame({"col": ["val"]})
        xlsx_config = {"file_format": "xlsx"}
        xlsx_reader = FileReader(xlsx_config)

        xlsx_reader.read_file("test.xlsx")

        mock_read_excel.assert_called_once()
        call_kwargs = mock_read_excel.call_args
        self.assertEqual(call_kwargs[1]["engine"], "openpyxl")

    @patch("drf_commons.services.import_from_file.core.file_reader.pd.read_excel")
    def test_read_file_strips_whitespace_from_headers(self, mock_read_excel):
        """Test that whitespace is stripped from column headers after reading."""
        mock_read_excel.return_value = pd.DataFrame({" username ": ["val"], " email": ["e@e.com"]})
        xlsx_config = {"file_format": "xlsx"}
        xlsx_reader = FileReader(xlsx_config)

        df = xlsx_reader.read_file("test.xlsx")

        self.assertIn("username", df.columns)
        self.assertIn("email", df.columns)

    def test_validate_headers_raises_on_missing_columns(self):
        """Test validate_headers raises ImportValidationError when columns are missing."""
        from drf_commons.services.import_from_file.core.exceptions import ImportValidationError

        required = {"username", "email", "first_name"}
        file_cols = ["username", "email"]

        with self.assertRaises(ImportValidationError) as cm:
            self.reader.validate_headers(file_cols, required)

        self.assertIn("Missing columns", str(cm.exception))
        self.assertIn("first_name", str(cm.exception))

    def test_validate_headers_raises_on_extra_columns(self):
        """Test validate_headers raises ImportValidationError when extra columns exist."""
        from drf_commons.services.import_from_file.core.exceptions import ImportValidationError

        required = {"username"}
        file_cols = ["username", "unexpected_col"]

        with self.assertRaises(ImportValidationError) as cm:
            self.reader.validate_headers(file_cols, required)

        self.assertIn("Unexpected columns", str(cm.exception))
        self.assertIn("unexpected_col", str(cm.exception))

    def test_validate_headers_passes_on_exact_match(self):
        """Test validate_headers passes when columns match exactly."""
        required = {"username", "email"}
        file_cols = ["username", "email"]

        # Should not raise
        self.reader.validate_headers(file_cols, required)

    def test_validate_headers_raises_on_both_missing_and_extra(self):
        """Test validate_headers raises on missing columns (checked before extra)."""
        from drf_commons.services.import_from_file.core.exceptions import ImportValidationError

        required = {"username", "email"}
        file_cols = ["username", "phone"]  # 'email' missing, 'phone' extra

        with self.assertRaises(ImportValidationError):
            self.reader.validate_headers(file_cols, required)

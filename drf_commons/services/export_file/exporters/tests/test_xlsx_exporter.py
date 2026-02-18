"""
Tests for Excel export functionality.

Tests Excel exporter functionality for exporting data to XLSX format.
"""

from io import BytesIO

from django.http import HttpResponse
from openpyxl import load_workbook

from drf_commons.common_tests.base_cases import DrfCommonTestCase

from ..xlsx_exporter import XLSXExporter


class XLSXExporterTests(DrfCommonTestCase):
    """Tests for XLSXExporter."""

    def setUp(self):
        super().setUp()
        self.exporter = XLSXExporter()
        self.sample_data = [
            {"id": 1, "name": "John Doe", "email": "john@example.com"},
            {"id": 2, "name": "Jane Smith", "email": "jane@example.com"},
        ]
        self.includes = ["id", "name", "email"]
        self.column_config = {
            "id": {"label": "ID"},
            "name": {"label": "Name"},
            "email": {"label": "Email"},
        }
        self.filename = "test_export.xlsx"
        self.export_headers = ["Test Export Report"]
        self.document_titles = ["User Data Export"]

    def test_export_returns_http_response(self):
        """Test export returns HttpResponse."""
        response = self.exporter.export(
            self.sample_data,
            self.includes,
            self.column_config,
            self.filename,
            self.export_headers,
            self.document_titles,
        )

        self.assertIsInstance(response, HttpResponse)
        expected_content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertEqual(response["Content-Type"], expected_content_type)
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn(self.filename, response["Content-Disposition"])

    def test_export_with_empty_data(self):
        """Test export with empty data returns response."""
        response = self.exporter.export(
            [],
            self.includes,
            self.column_config,
            self.filename,
            self.export_headers,
            self.document_titles,
        )

        self.assertIsInstance(response, HttpResponse)
        expected_content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertEqual(response["Content-Type"], expected_content_type)

    def test_export_content_type(self):
        """Test export sets correct content type."""
        response = self.exporter.export(
            self.sample_data,
            self.includes,
            self.column_config,
            self.filename,
            self.export_headers,
            self.document_titles,
        )

        expected_content_type = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertEqual(response["Content-Type"], expected_content_type)

    def test_export_filename_header(self):
        """Test export sets correct filename in headers."""
        response = self.exporter.export(
            self.sample_data,
            self.includes,
            self.column_config,
            self.filename,
            self.export_headers,
            self.document_titles,
        )

        content_disposition = response["Content-Disposition"]
        self.assertIn("attachment", content_disposition)
        self.assertIn(self.filename, content_disposition)

    def test_title_merge_handles_wide_columns(self):
        """Title merge must support excel columns beyond Z."""
        row = {f"col_{i}": f"value_{i}" for i in range(1, 54)}
        includes = list(row.keys())
        column_config = {field: {"label": field.upper()} for field in includes}

        response = self.exporter.export(
            [row],
            includes,
            column_config,
            self.filename,
            self.export_headers,
            ["Wide Export"],
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        merged_ranges = {str(cell_range) for cell_range in sheet.merged_cells.ranges}
        # Header writes row 1 then spacing row 2, so first title row is row 3.
        self.assertIn("A3:BA3", merged_ranges)

    def test_title_merge_boundaries_az_and_ba(self):
        """Merge ranges should remain valid at 52 (AZ) and 53 (BA) columns."""
        for total_columns, expected_last_col in [(52, "AZ"), (53, "BA")]:
            with self.subTest(total_columns=total_columns):
                row = {
                    f"field_{index}": f"value_{index}"
                    for index in range(1, total_columns + 1)
                }
                includes = list(row.keys())
                column_config = {
                    field_name: {"label": field_name}
                    for field_name in includes
                }

                response = self.exporter.export(
                    [row],
                    includes,
                    column_config,
                    self.filename,
                    self.export_headers,
                    ["Boundary Export"],
                )
                workbook = load_workbook(filename=BytesIO(response.content))
                sheet = workbook.active
                merged_ranges = {
                    str(cell_range) for cell_range in sheet.merged_cells.ranges
                }

                self.assertIn(f"A3:{expected_last_col}3", merged_ranges)

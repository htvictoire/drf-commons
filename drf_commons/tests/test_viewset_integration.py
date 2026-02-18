"""
ViewSet integration tests with real API calls.

Tests ViewSet mixins through DRF test client without mocking.
"""

import json
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.urls import path, include
from django.test import override_settings

from rest_framework import serializers, viewsets
from rest_framework.routers import DefaultRouter
from rest_framework.test import APITestCase

import openpyxl

from drf_commons.common_tests.factories import UserFactory
from drf_commons.serializers.base import BaseModelSerializer
from drf_commons.views.mixins import (
    BulkCreateModelMixin,
    BulkUpdateModelMixin,
    BulkDeleteModelMixin,
    FileImportMixin,
    FileExportMixin,
)
from drf_commons.views.mixins.crud import ListModelMixin

User = get_user_model()


class UserSerializer(serializers.ModelSerializer):
    """Test serializer for User model."""

    class Meta:
        model = User
        fields = ["id", "username", "email", "first_name", "last_name"]


class StandardUserViewSet(viewsets.ModelViewSet):
    """Standard ViewSet for individual CRUD operations."""

    queryset = User.objects.all()
    serializer_class = UserSerializer


class UserBulkSerializer(BaseModelSerializer):
    """Bulk serializer for User model with optimized operations."""

    class Meta(BaseModelSerializer.Meta):
        model = User
        fields = ["id", "username", "email", "first_name", "last_name"]


class BulkUserViewSet(
    viewsets.GenericViewSet,
    ListModelMixin,
    BulkCreateModelMixin,
    BulkUpdateModelMixin,
    BulkDeleteModelMixin,
):
    """ViewSet for bulk operations only."""

    queryset = User.objects.all()
    serializer_class = UserBulkSerializer


class ImportExportUserViewSet(
    viewsets.GenericViewSet,
    ListModelMixin,
    FileImportMixin,
    FileExportMixin,
):
    """ViewSet for import/export operations."""

    queryset = User.objects.all()
    serializer_class = UserSerializer

    # FileImportMixin configuration
    import_template_name = "user_import_template.xlsx"
    import_file_config = {
        "file_format": "xlsx",
        "order": ["users"],
        "models": {
            "users": {
                "model": "auth.User",
                "unique_by": ["username"],
                "update_if_exists": True,
                "direct_columns": {
                    "username": "username",
                    "email": "email",
                    "first_name": "first_name",
                    "last_name": "last_name"
                },
            }
        },
    }


# Test URLs configuration
router = DefaultRouter()
router.register(r'standard-users', StandardUserViewSet, basename='standard-user')
router.register(r'bulk-users', BulkUserViewSet, basename='bulk-user')
router.register(r'import-export-users', ImportExportUserViewSet, basename='import-export-user')

test_urlpatterns = [
    path('api/', include(router.urls)),
]


@override_settings(ROOT_URLCONF=__name__)
class StandardCRUDTests(APITestCase):
    """Test standard CRUD operations."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def test_standard_crud_operations(self):
        """Test standard CRUD operations work through API."""
        # CREATE
        create_data = {
            "username": "api_test_user",
            "email": "api@test.com",
            "first_name": "API",
            "last_name": "Test"
        }
        create_response = self.client.post('/api/standard-users/', create_data)
        self.assertEqual(create_response.status_code, 201)
        created_user_id = create_response.data['id']

        # READ (List)
        list_response = self.client.get('/api/standard-users/')
        self.assertEqual(list_response.status_code, 200)
        self.assertGreaterEqual(len(list_response.data), 1)

        # READ (Detail)
        detail_response = self.client.get(f'/api/standard-users/{created_user_id}/')
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.data['username'], 'api_test_user')

        # UPDATE
        update_data = {"email": "updated@test.com"}
        update_response = self.client.patch(f'/api/standard-users/{created_user_id}/', update_data)
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data['email'], 'updated@test.com')

        # DELETE
        delete_response = self.client.delete(f'/api/standard-users/{created_user_id}/')
        self.assertEqual(delete_response.status_code, 204)

        # Verify deletion
        verify_response = self.client.get(f'/api/standard-users/{created_user_id}/')
        self.assertEqual(verify_response.status_code, 404)


@override_settings(ROOT_URLCONF=__name__)
class BulkOperationTests(APITestCase):
    """Test bulk operations."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def test_bulk_create_operation(self):
        """Test bulk create operation through API."""
        bulk_data = [
            {
                "username": "bulk_user1",
                "email": "bulk1@test.com",
                "first_name": "Bulk",
                "last_name": "One"
            },
            {
                "username": "bulk_user2",
                "email": "bulk2@test.com",
                "first_name": "Bulk",
                "last_name": "Two"
            }
        ]

        response = self.client.post('/api/bulk-users/bulk-create/', bulk_data, format='json')
        self.assertEqual(response.status_code, 201)

        # Verify users were created in database
        self.assertTrue(User.objects.filter(username="bulk_user1").exists())
        self.assertTrue(User.objects.filter(username="bulk_user2").exists())

    def test_bulk_update_operation(self):
        """Test partial bulk update operation through API."""
        # Create test users
        user1 = UserFactory(username="update_user1", email="old1@test.com")
        user2 = UserFactory(username="update_user2", email="old2@test.com")

        bulk_update_data = [
            {"id": user1.id, "email": "new1@test.com"},
            {"id": user2.id, "email": "new2@test.com"}
        ]

        response = self.client.patch('/api/bulk-users/bulk-update/', bulk_update_data, format='json')
        self.assertEqual(response.status_code, 200)

        # Verify updates in database
        user1.refresh_from_db()
        user2.refresh_from_db()
        self.assertEqual(user1.email, "new1@test.com")
        self.assertEqual(user2.email, "new2@test.com")

    def test_bulk_update_matches_rows_by_id_not_queryset_order(self):
        """Bulk update must apply each row to its declared id, not positional queryset order."""
        user1 = UserFactory(username="ordered_user1", email="ordered_old1@test.com")
        user2 = UserFactory(username="ordered_user2", email="ordered_old2@test.com")

        # Reverse payload order intentionally. Querysets commonly return ascending PK order.
        bulk_update_data = [
            {"id": user2.id, "email": "ordered_new2@test.com"},
            {"id": user1.id, "email": "ordered_new1@test.com"},
        ]

        response = self.client.patch(
            "/api/bulk-users/bulk-update/", bulk_update_data, format="json"
        )
        self.assertEqual(response.status_code, 200)

        user1.refresh_from_db()
        user2.refresh_from_db()
        self.assertEqual(user1.email, "ordered_new1@test.com")
        self.assertEqual(user2.email, "ordered_new2@test.com")

    def test_bulk_put_requires_full_payload(self):
        """PUT bulk update must enforce full-update validation for each row."""
        user1 = UserFactory(username="put_user1", email="put_old1@test.com")
        user2 = UserFactory(username="put_user2", email="put_old2@test.com")

        # Missing required username fields for full update.
        payload = [
            {"id": user1.id, "email": "put_new1@test.com"},
            {"id": user2.id, "email": "put_new2@test.com"},
        ]

        response = self.client.put("/api/bulk-users/bulk-update/", payload, format="json")
        self.assertEqual(response.status_code, 400)

        user1.refresh_from_db()
        user2.refresh_from_db()
        self.assertEqual(user1.email, "put_old1@test.com")
        self.assertEqual(user2.email, "put_old2@test.com")

    def test_bulk_update_rejects_rows_with_missing_or_inaccessible_ids(self):
        """Bulk update should fail fast when any row id is missing from queryset."""
        user1 = UserFactory(username="missing_id_user1", email="missing_old1@test.com")

        payload = [
            {"id": user1.id, "email": "missing_new1@test.com"},
            {"id": 99999999, "email": "missing_new_unknown@test.com"},
        ]

        response = self.client.put("/api/bulk-users/bulk-update/", payload, format="json")
        self.assertEqual(response.status_code, 400)

        user1.refresh_from_db()
        self.assertEqual(user1.email, "missing_old1@test.com")

    def test_bulk_update_rejects_duplicate_ids(self):
        """Bulk update should reject duplicate ids to prevent ambiguous row mapping."""
        user1 = UserFactory(username="dup_user1", email="dup_old1@test.com")

        payload = [
            {"id": user1.id, "email": "dup_new1@test.com"},
            {"id": user1.id, "email": "dup_new2@test.com"},
        ]

        response = self.client.put("/api/bulk-users/bulk-update/", payload, format="json")
        self.assertEqual(response.status_code, 400)

        user1.refresh_from_db()
        self.assertEqual(user1.email, "dup_old1@test.com")

    def test_bulk_delete_operation(self):
        """Test bulk delete operation through API."""
        # Create test users
        user1 = UserFactory(username="delete_user1")
        user2 = UserFactory(username="delete_user2")
        user3 = UserFactory(username="keep_user")

        delete_ids = [user1.id, user2.id]
        response = self.client.delete('/api/bulk-users/bulk_delete/', delete_ids, format='json')
        self.assertEqual(response.status_code, 200)

        # Verify deletions in database
        self.assertFalse(User.objects.filter(id=user1.id).exists())
        self.assertFalse(User.objects.filter(id=user2.id).exists())
        self.assertTrue(User.objects.filter(id=user3.id).exists())

    def test_bulk_delete_count_excludes_cascaded_rows(self):
        """Bulk delete response count should report only directly deleted model rows."""
        user1 = UserFactory(username="cascade_delete_user1")
        user2 = UserFactory(username="cascade_delete_user2")
        group = Group.objects.create(name="cascade_delete_group")
        user1.groups.add(group)
        user2.groups.add(group)

        through_model = User.groups.through
        self.assertEqual(through_model.objects.filter(group=group).count(), 2)

        delete_ids = [user1.id, user2.id]
        response = self.client.delete("/api/bulk-users/bulk_delete/", delete_ids, format="json")
        self.assertEqual(response.status_code, 200)

        self.assertEqual(response.data["data"]["count"], 2)
        self.assertEqual(response.data["data"]["requested_count"], 2)
        self.assertEqual(through_model.objects.filter(group=group).count(), 0)


@override_settings(ROOT_URLCONF=__name__)
class ImportExportTests(APITestCase):
    """Test import/export operations."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def create_test_excel_file(self, data_rows):
        """Create Excel file for import testing."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "users"

        # Add title rows (service expects header on row 5)
        sheet.cell(row=1, column=1, value="User Import Template")
        sheet.cell(row=2, column=1, value="Fill in the data below")
        sheet.cell(row=3, column=1, value="")
        sheet.cell(row=4, column=1, value="")

        # Headers on row 5
        headers = ["username", "email", "first_name", "last_name"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=5, column=col, value=header)

        # Data rows starting from row 6
        for row_idx, row_data in enumerate(data_rows, 6):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)
        return file_buffer

    def test_file_import_operation(self):
        """Test file import operation through API."""
        test_data = [
            ["import_user1", "import1@test.com", "Import", "One"],
            ["import_user2", "import2@test.com", "Import", "Two"],
        ]

        excel_file = self.create_test_excel_file(test_data)

        response = self.client.post(
            '/api/import-export-users/import-from-file/',
            {
                'file': excel_file,
                'append_data': 'true',
                'config': json.dumps({
                    "file_format": "xlsx",
                    "order": ["users"],
                    "models": {
                        "users": {
                            "model": "auth.User",
                            "unique_by": ["username"],
                            "update_if_exists": False,
                            "direct_columns": {
                                "username": "username",
                                "email": "email",
                                "first_name": "first_name",
                                "last_name": "last_name"
                            },
                        }
                    },
                })
            },
            format='multipart'
        )
        self.assertIn(response.status_code, [200, 201])  # Accept both 200 and 201 for successful import

        # Check the import summary from the response
        if hasattr(response, 'data'):
            if 'data' in response.data and 'import_summary' in response.data['data']:
                summary = response.data['data']['import_summary']
                self.assertEqual(summary['created'], 2)
                self.assertEqual(summary['failed'], 0)
            elif 'successful_rows' in response.data:
                self.assertEqual(response.data['successful_rows'], 2)
                self.assertEqual(len(response.data['errors']), 0)

        # Verify users were imported
        self.assertTrue(User.objects.filter(username="import_user1").exists())
        self.assertTrue(User.objects.filter(username="import_user2").exists())

    def test_file_import_operation_accepts_boolean_mode_flag(self):
        """Import endpoint should accept boolean mode flags without crashing."""
        test_data = [
            ["import_bool_user", "import_bool@test.com", "Import", "Bool"],
        ]
        excel_file = self.create_test_excel_file(test_data)

        response = self.client.post(
            "/api/import-export-users/import-from-file/",
            {"file": excel_file, "append_data": True},
            format="multipart",
        )
        self.assertIn(response.status_code, [200, 201])
        self.assertTrue(User.objects.filter(username="import_bool_user").exists())

    def test_file_import_updates_same_file_duplicates_when_update_enabled(self):
        """Duplicate unique_by keys in the same file should resolve to one created record plus updates."""
        duplicate_key_data = [
            ["dup_in_file_user", "first@test.com", "First", "Version"],
            ["dup_in_file_user", "second@test.com", "Second", "Version"],
        ]
        excel_file = self.create_test_excel_file(duplicate_key_data)

        response = self.client.post(
            "/api/import-export-users/import-from-file/",
            {
                "file": excel_file,
                "append_data": "true",
                "config": json.dumps(
                    {
                        "file_format": "xlsx",
                        "order": ["users"],
                        "models": {
                            "users": {
                                "model": "auth.User",
                                "unique_by": ["username"],
                                "update_if_exists": True,
                                "direct_columns": {
                                    "username": "username",
                                    "email": "email",
                                    "first_name": "first_name",
                                    "last_name": "last_name",
                                },
                            }
                        },
                    }
                ),
            },
            format="multipart",
        )
        self.assertIn(response.status_code, [200, 201])

        user_qs = User.objects.filter(username="dup_in_file_user")
        self.assertEqual(user_qs.count(), 1)
        user = user_qs.first()
        self.assertEqual(user.email, "second@test.com")
        self.assertEqual(user.first_name, "Second")

        if hasattr(response, "data"):
            if "data" in response.data and "import_summary" in response.data["data"]:
                summary = response.data["data"]["import_summary"]
                self.assertEqual(summary["created"], 1)
                self.assertEqual(summary["updated"], 1)
                self.assertEqual(summary["failed"], 0)

    def test_file_import_replace_mode_replaces_data_on_full_success(self):
        """replace_data=true should atomically replace existing queryset rows on successful import."""
        UserFactory(username="replace_old_user", email="old_replace@test.com")

        test_data = [
            ["replace_new_user1", "replace1@test.com", "Replace", "One"],
            ["replace_new_user2", "replace2@test.com", "Replace", "Two"],
        ]
        excel_file = self.create_test_excel_file(test_data)

        response = self.client.post(
            "/api/import-export-users/import-from-file/",
            {"file": excel_file, "replace_data": "true"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 201)

        self.assertFalse(User.objects.filter(username="replace_old_user").exists())
        self.assertTrue(User.objects.filter(username="replace_new_user1").exists())
        self.assertTrue(User.objects.filter(username="replace_new_user2").exists())

        self.assertTrue(response.data.get("success"))
        self.assertEqual(response.data["data"]["operation"], "replace")
        self.assertGreaterEqual(response.data["data"]["deleted_count"], 1)

    def test_file_import_replace_mode_rolls_back_when_any_row_fails(self):
        """replace_data=true should roll back delete+import if any imported row fails."""
        UserFactory(username="replace_keep_user", email="keep_replace@test.com")

        # Missing required username forces a row-level create failure.
        failing_data = [
            ["replace_valid_user", "valid@test.com", "Valid", "Row"],
            [None, "invalid@test.com", "Invalid", "Row"],
        ]
        excel_file = self.create_test_excel_file(failing_data)

        response = self.client.post(
            "/api/import-export-users/import-from-file/",
            {"file": excel_file, "replace_data": "true"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 422)

        # Existing data must remain because replace operation is atomic.
        self.assertTrue(User.objects.filter(username="replace_keep_user").exists())
        # New rows from failed import must not be committed.
        self.assertFalse(User.objects.filter(username="replace_valid_user").exists())

        self.assertFalse(response.data.get("success"))
        self.assertEqual(response.data["data"]["operation"], "replace")
        self.assertEqual(response.data["data"]["deleted_count"], 0)
        self.assertGreater(response.data["data"]["import_summary"]["failed"], 0)

    def test_file_export_csv_operation(self):
        """Test CSV export operation through API."""
        # Create test data
        user1 = UserFactory(username="export_user1", email="export1@test.com", first_name="Export")
        user2 = UserFactory(username="export_user2", email="export2@test.com", first_name="Test")

        export_data = {
            "file_type": "csv",
            "includes": ["username", "email", "first_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"}
            },
            "data": [
                {"username": user1.username, "email": user1.email, "first_name": user1.first_name},
                {"username": user2.username, "email": user2.email, "first_name": user2.first_name}
            ]
        }

        response = self.client.post('/api/import-export-users/export-as-file/', export_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')

        # Verify CSV content
        content = response.content.decode('utf-8')
        self.assertIn("Username,Email,First Name", content)
        self.assertIn("export_user1,export1@test.com,Export", content)

    def test_file_export_csv_operation_with_comma_separated_includes(self):
        """Comma-separated includes payload should be normalized to field list."""
        user = UserFactory(username="csv_inc_user", email="csv_inc@test.com", first_name="Csv")

        export_data = {
            "file_type": "csv",
            "includes": " username, email , first_name ",
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"},
            },
            "data": [
                {"username": user.username, "email": user.email, "first_name": user.first_name}
            ],
        }

        response = self.client.post('/api/import-export-users/export-as-file/', export_data)
        self.assertEqual(response.status_code, 200)

        content = response.content.decode("utf-8")
        self.assertIn("Username,Email,First Name", content)
        self.assertIn("csv_inc_user,csv_inc@test.com,Csv", content)

    def test_file_export_rejects_invalid_includes_type(self):
        """Non-string/non-list includes payload should be rejected."""
        user = UserFactory(username="bad_inc_user", email="bad_inc@test.com", first_name="Bad")

        export_data = {
            "file_type": "csv",
            "includes": 123,
            "column_config": {
                "username": {"label": "Username"},
            },
            "data": [
                {"username": user.username},
            ],
        }

        response = self.client.post('/api/import-export-users/export-as-file/', export_data)
        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.data["success"])
        self.assertIn("includes", response.data["errors"])

    def test_file_export_xlsx_operation(self):
        """Test Excel export operation through API."""
        user = UserFactory(username="xlsx_user", email="xlsx@test.com", first_name="Excel")

        export_data = {
            "file_type": "xlsx",
            "includes": ["username", "email", "first_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"}
            },
            "data": [
                {"username": user.username, "email": user.email, "first_name": user.first_name}
            ]
        }

        response = self.client.post('/api/import-export-users/export-as-file/', export_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Verify Excel content can be loaded
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertIsNotNone(sheet.cell(row=1, column=1).value)  # Has content


@override_settings(ROOT_URLCONF=__name__)
class CombinedWorkflowTests(APITestCase):
    """Test workflows combining different operation types."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def test_combined_operations_workflow(self):
        """Test workflow combining bulk and export operations."""
        # 1. Bulk create initial data
        initial_data = [
            {"username": "workflow1", "email": "flow1@test.com", "first_name": "Work"},
            {"username": "workflow2", "email": "flow2@test.com", "first_name": "Flow"}
        ]
        create_response = self.client.post('/api/bulk-users/bulk-create/', initial_data, format='json')
        self.assertEqual(create_response.status_code, 201)

        # 2. List to verify creation
        list_response = self.client.get('/api/bulk-users/')
        # Extract users from the response data structure
        response_data = list_response.data.get('data', {})
        results = response_data.get('results', [])
        workflow_users = [
            user for user in results
            if user['username'].startswith('workflow')
        ]
        self.assertEqual(len(workflow_users), 2)

        # 3. Bulk update the created users
        update_data = [
            {"id": workflow_users[0]['id'], "last_name": "Updated"},
            {"id": workflow_users[1]['id'], "last_name": "Updated"}
        ]
        update_response = self.client.patch('/api/bulk-users/bulk-update/', update_data, format='json')
        self.assertEqual(update_response.status_code, 200)

        # 4. Export the updated data
        export_data = [
            {
                "username": workflow_users[0]['username'],
                "email": workflow_users[0]['email'],
                "first_name": workflow_users[0]['first_name'],
                "last_name": "Updated"
            },
            {
                "username": workflow_users[1]['username'],
                "email": workflow_users[1]['email'],
                "first_name": workflow_users[1]['first_name'],
                "last_name": "Updated"
            }
        ]
        export_response = self.client.post('/api/import-export-users/export-as-file/', {
            "file_type": "csv",
            "includes": ["username", "email", "first_name", "last_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"},
                "last_name": {"label": "Last Name"}
            },
            "data": export_data
        })
        self.assertEqual(export_response.status_code, 200)

        # Verify export contains updated data
        content = export_response.content.decode('utf-8')
        self.assertIn("workflow1", content)
        self.assertIn("Updated", content)

        # 5. Clean up with bulk delete
        delete_ids = [user['id'] for user in workflow_users]
        delete_response = self.client.delete('/api/bulk-users/bulk_delete/', delete_ids, format='json')
        self.assertEqual(delete_response.status_code, 200)

        # Verify cleanup
        final_list = self.client.get('/api/bulk-users/')
        final_response_data = final_list.data.get('data', {})
        final_results = final_response_data.get('results', [])
        final_workflow_users = [
            user for user in final_results
            if user['username'].startswith('workflow')
        ]
        self.assertEqual(len(final_workflow_users), 0)


# Make this module act as a URLconf for testing
urlpatterns = test_urlpatterns
